from flask import Flask, request, jsonify, send_from_directory, session, redirect, Response
from flask_cors import CORS
import sqlite3, os, io
from datetime import datetime
from functools import wraps

try:
    from twilio.rest import Client as TwilioClient
    TWILIO_OK = True
except ImportError:
    TWILIO_OK = False

try:
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    EXCEL_OK = True
except ImportError:
    EXCEL_OK = False

app = Flask(__name__, static_folder='static', static_url_path='')
app.secret_key = os.environ.get('SECRET_KEY', 'klinik-gizli-2024')
CORS(app, supports_credentials=True)

DB       = os.environ.get('DB_PATH', '/tmp/klinik.db')
KULLANICI = os.environ.get('KULLANICI', 'admin')
SIFRE    = os.environ.get('SIFRE', 'klinik123')
TWILIO_SID   = os.environ.get('TWILIO_SID', '')
TWILIO_TOKEN = os.environ.get('TWILIO_TOKEN', '')
TWILIO_FROM  = os.environ.get('TWILIO_FROM', '')
SMS_TO       = os.environ.get('SMS_TO', '')

def kon():
    c = sqlite3.connect(DB)
    c.row_factory = sqlite3.Row
    return c

def kurulum():
    db = kon()
    db.execute('''CREATE TABLE IF NOT EXISTS kayitlar (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        hasta TEXT NOT NULL,
        tarih TEXT NOT NULL,
        doktor TEXT NOT NULL,
        islemler TEXT NOT NULL,
        ucret REAL NOT NULL DEFAULT 0,
        odeme TEXT NOT NULL DEFAULT 'nakit',
        notlar TEXT DEFAULT ''
    )''')
    db.commit()
    db.close()

kurulum()

def giris_gerekli(f):
    @wraps(f)
    def kontrol(*args, **kwargs):
        if not session.get('giris'):
            if request.path.startswith('/api/'):
                return jsonify({'hata': 'Giriş gerekli'}), 401
            return redirect('/giris')
        return f(*args, **kwargs)
    return kontrol

# ── Auth ──────────────────────────────────────────────────
@app.route('/api/giris', methods=['POST'])
def giris():
    d = request.get_json()
    if d.get('kullanici') == KULLANICI and d.get('sifre') == SIFRE:
        session['giris'] = True
        return jsonify({'ok': True})
    return jsonify({'hata': 'Kullanıcı adı veya şifre hatalı'}), 401

@app.route('/api/cikis', methods=['POST'])
def cikis():
    session.clear()
    return jsonify({'ok': True})

@app.route('/api/durum')
def durum():
    return jsonify({'giris': bool(session.get('giris'))})

# ── Kayıtlar ──────────────────────────────────────────────
@app.route('/api/kayitlar', methods=['GET'])
@giris_gerekli
def listele():
    db = kon()
    rows = db.execute('SELECT * FROM kayitlar ORDER BY tarih DESC').fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/kayitlar', methods=['POST'])
@giris_gerekli
def ekle():
    d = request.get_json()
    if not d or not d.get('hasta') or not d.get('tarih'):
        return jsonify({'hata': 'Eksik bilgi'}), 400
    islemler = ','.join(d['islemler']) if isinstance(d['islemler'], list) else d['islemler']
    db = kon()
    cur = db.execute(
        'INSERT INTO kayitlar (hasta,tarih,doktor,islemler,ucret,odeme,notlar) VALUES (?,?,?,?,?,?,?)',
        (d['hasta'], d['tarih'], d.get('doktor','ortopedi'), islemler,
         float(d.get('ucret',0)), d.get('odeme','nakit'), d.get('notlar',''))
    )
    yeni_id = cur.lastrowid
    db.commit()
    row = db.execute('SELECT * FROM kayitlar WHERE id=?', (yeni_id,)).fetchone()
    db.close()
    return jsonify(dict(row)), 201

@app.route('/api/kayitlar/<int:kid>', methods=['PUT'])
@giris_gerekli
def guncelle(kid):
    d = request.get_json()
    islemler = ','.join(d['islemler']) if isinstance(d['islemler'], list) else d['islemler']
    db = kon()
    db.execute(
        'UPDATE kayitlar SET hasta=?,tarih=?,doktor=?,islemler=?,ucret=?,odeme=?,notlar=? WHERE id=?',
        (d['hasta'], d['tarih'], d.get('doktor','ortopedi'), islemler,
         float(d.get('ucret',0)), d.get('odeme','nakit'), d.get('notlar',''), kid)
    )
    db.commit()
    row = db.execute('SELECT * FROM kayitlar WHERE id=?', (kid,)).fetchone()
    db.close()
    return jsonify(dict(row))

@app.route('/api/kayitlar/<int:kid>', methods=['DELETE'])
@giris_gerekli
def sil(kid):
    db = kon()
    db.execute('DELETE FROM kayitlar WHERE id=?', (kid,))
    db.commit()
    db.close()
    return jsonify({'ok': True})

# ── Rapor ─────────────────────────────────────────────────
@app.route('/api/rapor')
@giris_gerekli
def rapor():
    bas = request.args.get('bas', '2000-01-01')
    bit = request.args.get('bit', '2099-12-31')
    doc = request.args.get('doktor', '')
    db = kon()
    q = 'SELECT * FROM kayitlar WHERE tarih>=? AND tarih<=?'
    p = [bas+'T00:00:00', bit+'T23:59:59']
    if doc:
        q += ' AND doktor=?'; p.append(doc)
    rows = [dict(r) for r in db.execute(q, p).fetchall()]
    db.close()
    return jsonify({
        'kayitlar': rows,
        'toplam': sum(r['ucret'] for r in rows),
        'nakit':  sum(r['ucret'] for r in rows if r['odeme']=='nakit'),
        'kk':     sum(r['ucret'] for r in rows if r['odeme']=='kk'),
        'havale': sum(r['ucret'] for r in rows if r['odeme']=='havale'),
    })

# ── Hasta Geçmişi ─────────────────────────────────────────
@app.route('/api/hastalar')
@giris_gerekli
def hastalar():
    db = kon()
    rows = db.execute('''
        SELECT hasta, COUNT(*) as ziyaret, SUM(ucret) as toplam, MAX(tarih) as son_tarih
        FROM kayitlar GROUP BY hasta ORDER BY son_tarih DESC
    ''').fetchall()
    db.close()
    return jsonify([dict(r) for r in rows])

@app.route('/api/hasta/<path:ad>')
@giris_gerekli
def hasta_gecmis(ad):
    db = kon()
    rows = [dict(r) for r in db.execute(
        'SELECT * FROM kayitlar WHERE hasta=? ORDER BY tarih DESC', (ad,)
    ).fetchall()]
    db.close()
    return jsonify({
        'kayitlar': rows,
        'toplam': sum(r['ucret'] for r in rows),
        'ziyaret': len(rows)
    })

# ── Aylık Karşılaştırma ───────────────────────────────────
@app.route('/api/aylik')
@giris_gerekli
def aylik():
    db = kon()
    ozet = db.execute('''
        SELECT strftime('%Y-%m', tarih) AS ay, doktor, odeme,
               COUNT(*) AS hasta_sayisi, SUM(ucret) AS toplam
        FROM kayitlar GROUP BY ay, doktor, odeme ORDER BY ay
    ''').fetchall()
    db.close()
    return jsonify([dict(r) for r in ozet])

# ── Excel Export ──────────────────────────────────────────
@app.route('/api/excel')
@giris_gerekli
def excel_indir():
    bas = request.args.get('bas', '2000-01-01')
    bit = request.args.get('bit', '2099-12-31')
    doc = request.args.get('doktor', '')
    db = kon()
    q = 'SELECT * FROM kayitlar WHERE tarih>=? AND tarih<=?'
    p = [bas+'T00:00:00', bit+'T23:59:59']
    if doc:
        q += ' AND doktor=?'; p.append(doc)
    rows = [dict(r) for r in db.execute(q, p).fetchall()]
    db.close()

    if not EXCEL_OK:
        return jsonify({'hata': 'openpyxl kurulu değil'}), 500

    wb = openpyxl.Workbook()

    # Sayfa 1 — Kayıtlar
    ws = wb.active
    ws.title = 'Kayıtlar'
    basliklar = ['ID','Hasta','Tarih','Doktor','İşlemler','Ücret (₺)','Ödeme','Notlar']
    for i, b in enumerate(basliklar, 1):
        h = ws.cell(1, i, b)
        h.fill = PatternFill('solid', fgColor='1255B8')
        h.font = Font(bold=True, color='FFFFFF')
        h.alignment = Alignment(horizontal='center')
    for ri, row in enumerate(rows, 2):
        ws.cell(ri, 1, row['id'])
        ws.cell(ri, 2, row['hasta'])
        ws.cell(ri, 3, row['tarih'][:16].replace('T',' '))
        ws.cell(ri, 4, 'Ortopedi' if row['doktor']=='ortopedi' else 'Fizik Tedavi')
        ws.cell(ri, 5, row['islemler'])
        c = ws.cell(ri, 6, row['ucret']); c.number_format = '#,##0'
        ws.cell(ri, 7, {'nakit':'Nakit','kk':'Kredi Kartı','havale':'Havale'}.get(row['odeme'], row['odeme']))
        ws.cell(ri, 8, row['notlar'])
    son = len(rows)+3
    ws.cell(son, 5, 'TOPLAM').font = Font(bold=True)
    c = ws.cell(son, 6, sum(r['ucret'] for r in rows))
    c.font = Font(bold=True); c.number_format = '#,##0'
    for col, w in zip('ABCDEFGH', [6,22,18,14,35,14,12,20]):
        ws.column_dimensions[col].width = w

    # Sayfa 2 — Aylık Özet
    ws2 = wb.create_sheet('Aylık Özet')
    for i, b in enumerate(['Ay','Hasta Sayısı','Toplam Gelir (₺)','Ortopedi','Fizik T.','Nakit','K.Kartı','Havale'], 1):
        h = ws2.cell(1, i, b)
        h.fill = PatternFill('solid', fgColor='1255B8')
        h.font = Font(bold=True, color='FFFFFF')
    ay_map = {}
    for row in rows:
        ay = row['tarih'][:7]
        if ay not in ay_map:
            ay_map[ay] = {'hasta':0,'toplam':0,'ort':0,'fiz':0,'nakit':0,'kk':0,'havale':0}
        ay_map[ay]['hasta'] += 1
        ay_map[ay]['toplam'] += row['ucret']
        ay_map[ay]['ort' if row['doktor']=='ortopedi' else 'fiz'] += row['ucret']
        ay_map[ay][row['odeme']] = ay_map[ay].get(row['odeme'], 0) + row['ucret']
    for ri, (ay, v) in enumerate(sorted(ay_map.items(), reverse=True), 2):
        ws2.cell(ri, 1, ay)
        ws2.cell(ri, 2, v['hasta'])
        for ci, key in enumerate(['toplam','ort','fiz','nakit','kk','havale'], 3):
            c = ws2.cell(ri, ci, v.get(key, 0)); c.number_format = '#,##0'
    for col, w in zip('ABCDEFGH', [10,14,18,14,14,12,12,12]):
        ws2.column_dimensions[col].width = w

    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return Response(buf.getvalue(),
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        headers={'Content-Disposition': f'attachment;filename=klinik_{bas}_{bit}.xlsx'})

# ── Günlük SMS Özet ───────────────────────────────────────
@app.route('/api/sms', methods=['POST'])
@giris_gerekli
def sms_gonder():
    if not TWILIO_OK:
        return jsonify({'hata': 'Twilio kurulu değil'}), 400
    if not all([TWILIO_SID, TWILIO_TOKEN, TWILIO_FROM, SMS_TO]):
        return jsonify({'hata': 'SMS ayarları eksik. TWILIO_SID, TWILIO_TOKEN, TWILIO_FROM, SMS_TO değişkenlerini Railway Variables bölümüne ekleyin.'}), 400
    bugun = datetime.now().strftime('%Y-%m-%d')
    db = kon()
    rows = [dict(r) for r in db.execute(
        'SELECT * FROM kayitlar WHERE tarih>=? AND tarih<=?',
        [bugun+'T00:00:00', bugun+'T23:59:59']
    ).fetchall()]
    db.close()
    if not rows:
        mesaj = f'Klinik {bugun}: Bugün kayıt yok.'
    else:
        toplam = sum(r['ucret'] for r in rows)
        nakit  = sum(r['ucret'] for r in rows if r['odeme']=='nakit')
        kk     = sum(r['ucret'] for r in rows if r['odeme']=='kk')
        havale = sum(r['ucret'] for r in rows if r['odeme']=='havale')
        ort = len([r for r in rows if r['doktor']=='ortopedi'])
        fiz = len([r for r in rows if r['doktor']=='fizik'])
        mesaj = (f'Klinik {bugun}\n'
                 f'Hasta:{len(rows)} (Ort:{ort} Fiz:{fiz})\n'
                 f'Toplam:{int(toplam):,}TL\n'
                 f'Nakit:{int(nakit):,} KK:{int(kk):,} Havale:{int(havale):,}')
    try:
        TwilioClient(TWILIO_SID, TWILIO_TOKEN).messages.create(body=mesaj, from_=TWILIO_FROM, to=SMS_TO)
        return jsonify({'ok': True, 'mesaj': mesaj})
    except Exception as e:
        return jsonify({'hata': str(e)}), 500

# ── SPA ───────────────────────────────────────────────────
@app.route('/')
def ana():
    if not session.get('giris'):
        return send_from_directory(app.static_folder, 'giris.html')
    return send_from_directory(app.static_folder, 'index.html')

@app.route('/giris')
def giris_route():
    return send_from_directory(app.static_folder, 'giris.html')

@app.route('/<path:path>')
def spa(path):
    dosya = os.path.join(app.static_folder, path)
    if os.path.exists(dosya):
        return send_from_directory(app.static_folder, path)
    if not session.get('giris'):
        return send_from_directory(app.static_folder, 'giris.html')
    return send_from_directory(app.static_folder, 'index.html')
if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
